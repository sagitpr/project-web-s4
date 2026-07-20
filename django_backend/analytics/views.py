"""
Analytics views for Warungio Marketplace.
Real-time seller dashboard analytics.
"""

from django.db.models import Sum, Count, Avg, Q, Min
from django.db.models.functions import TruncDate, TruncMonth, TruncDay
from django.utils import timezone
from datetime import timedelta, date, datetime
from decimal import Decimal
from rest_framework import status, generics, permissions, views
from rest_framework.response import Response
from django.core.cache import cache

from .models import SalesAnalytics, DeviceAnalytics, UserActivity, DailyReport
from .serializers import (
    SalesAnalyticsSerializer, SalesSummarySerializer, SalesTrendSerializer,
    DeviceAnalyticsSerializer, DeviceBreakdownSerializer,
    UserActivitySerializer, DashboardStatsSerializer
)
from orders.models import Order, OrderItem
from products.models import Product, Review
from stores.models import StoreFollower
from accounts.permissions import IsSeller, IsAdmin
from .services.ai_insight import AISellerInsightService
from drf_spectacular.utils import extend_schema


@extend_schema(exclude=True)
class DashboardSummaryView(views.APIView):
    """Get seller dashboard summary statistics (cached 2 min)."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        store = request.user.store
        period = request.query_params.get('period', 'month')  # week, month, year

        # Cache key includes store + period
        cache_key = f'dashboard_summary_{store.id}_{period}'
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        # Calculate date range
        today = timezone.now().date()
        if period == 'week':
            start_date = today - timedelta(days=7)
        elif period == 'month':
            start_date = today - timedelta(days=30)
        elif period == 'year':
            start_date = today - timedelta(days=365)
        else:
            start_date = today - timedelta(days=30)

        # Previous period for trend comparison
        period_days = (today - start_date).days or 1
        prev_start = start_date - timedelta(days=period_days)
        prev_end = start_date - timedelta(days=1)
        
        # Orders in period (completed/paid)
        orders_qs = Order.objects.filter(
            store=store,
            created_at__date__gte=start_date,
            order_status__in=['paid', 'processed', 'shipped', 'completed']
        )
        # Previous period orders
        prev_orders_qs = Order.objects.filter(
            store=store,
            created_at__date__gte=prev_start,
            created_at__date__lte=prev_end,
            order_status__in=['paid', 'processed', 'shipped', 'completed']
        )
        
        # Calculate stats
        total_sales = orders_qs.aggregate(
            total=Sum('total_price')
        )['total'] or 0
        prev_total_sales = prev_orders_qs.aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        total_orders = orders_qs.count()
        prev_total_orders = prev_orders_qs.count()
        
        # Trend percentages
        def calc_pct(current, previous):
            if previous == 0:
                return 100.0 if current > 0 else 0.0
            return round(((float(current) - float(previous)) / float(previous)) * 100, 1)
        
        sales_trend_percent = calc_pct(total_sales, prev_total_sales)
        order_trend_percent = calc_pct(total_orders, prev_total_orders)
        
        # Products sold
        products_sold = OrderItem.objects.filter(
            order__in=orders_qs
        ).aggregate(total=Sum('qty'))['total'] or 0
        
        # New customers (store followers in period)
        new_customers = StoreFollower.objects.filter(
            store=store,
            created_at__date__gte=start_date
        ).count()
        
        # Total followers
        total_followers = StoreFollower.objects.filter(store=store).count()
        
        # Average rating
        avg_rating = Review.objects.filter(
            product__store=store
        ).aggregate(avg=Avg('rating'))['avg'] or 0
        total_reviews = Review.objects.filter(
            product__store=store
        ).count()
        
        # Total products
        total_products = Product.objects.filter(store=store, is_active=True).count()
        
        # Pending orders
        pending_orders = Order.objects.filter(
            store=store, order_status='pending'
        ).count()
        
        # Quality summary — single annotated query instead of 4-8 separate queries
        qs = Product.objects.filter(store=store, is_active=True)
        quality_counts = qs.aggregate(
            fresh=Count('id', filter=Q(product_status='fresh', quality_score__gte=80)),
            normal=Count('id', filter=Q(product_status='normal', quality_score__gte=60) & ~Q(product_status='fresh')),
            warning=Count('id', filter=Q(quality_score__lt=60, quality_score__gte=30)),
            rejected=Count('id', filter=Q(quality_score__lt=30) & ~Q(quality_score=0)),
        )
        fresh_count = quality_counts['fresh']
        normal_count = quality_counts['normal']
        warning_count = quality_counts['warning']
        rejected_count = quality_counts['rejected']
        if fresh_count + normal_count + warning_count + rejected_count == 0:
            # Fallback to product_status if quality_score not set — single annotated query
            fallback = qs.aggregate(
                fresh=Count('id', filter=Q(product_status='fresh')),
                normal=Count('id', filter=Q(product_status='normal') & ~Q(product_status='fresh')),
                warning=Count('id', filter=Q(product_status='low')),
                rejected=Count('id', filter=Q(product_status='bad')),
            )
            fresh_count = fallback['fresh']
            normal_count = fallback['normal']
            warning_count = fallback['warning']
            rejected_count = fallback['rejected']
        
        # Sales trend data
        sales_data = self._get_sales_trend(store, start_date, today)
        
        # Device breakdown
        device_data = self._get_device_breakdown(store, start_date)
        
        # Recent orders
        recent_orders = self._get_recent_orders(store)
        
        # Top products
        top_products = self._get_top_products(store, start_date)
        
        data = {
            'total_sales': total_sales,
            'total_orders': total_orders,
            'products_sold': products_sold,
            'new_customers': new_customers,
            'total_followers': total_followers,
            'average_rating': round(avg_rating, 2),
            'total_reviews': total_reviews,
            'total_products': total_products,
            'pending_orders': pending_orders,
            'sales_trend_percent': sales_trend_percent,
            'order_trend_percent': order_trend_percent,
            'quality_summary': {
                'fresh': fresh_count,
                'normal': normal_count,
                'warning': warning_count,
                'rejected': rejected_count,
            },
            'sales_chart': sales_data,
            'device_breakdown': device_data,
            'recent_orders': recent_orders,
            'top_products': top_products,
        }

        # Cache briefly for spike protection — dashboard always shows near-live data
        cache.set(cache_key, data, 15)
        return Response(data)

    def _get_sales_trend(self, store, start_date, end_date):
        """Get daily sales data for charts."""
        orders = Order.objects.filter(
            store=store,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            order_status__in=['paid', 'processed', 'shipped', 'completed']
        ).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            daily_sales=Sum('total_price'),
            daily_orders=Count('id')
        ).order_by('day')
        
        # Build chart data — use dict indexing instead of O(n*m) linear scan
        indexed = {}
        for o in orders:
            if o['day']:
                indexed[o['day'].date()] = o
        
        days = []
        sales = []
        order_counts = []
        
        current = start_date
        while current <= end_date:
            days.append(current.strftime('%d %b'))
            found = indexed.get(current)
            if found:
                sales.append(float(found['daily_sales']))
                order_counts.append(found['daily_orders'])
            else:
                sales.append(0)
                order_counts.append(0)
            current += timedelta(days=1)
        
        return {
            'labels': days,
            'daily_sales': sales,
            'daily_orders': order_counts,
            'daily_customers': [],
        }

    def _get_device_breakdown(self, store, start_date):
        """Get device usage breakdown."""
        activities = UserActivity.objects.filter(
            store=store,
            created_at__date__gte=start_date
        )
        
        mobile = activities.filter(device_type='mobile').count()
        tablet = activities.filter(device_type='tablet').count()
        desktop = activities.filter(device_type='desktop').count()
        total = mobile + tablet + desktop or 1
        
        return {
            'mobile_percentage': round(mobile / total * 100, 2),
            'tablet_percentage': round(tablet / total * 100, 2),
            'desktop_percentage': round(desktop / total * 100, 2),
            'mobile_count': mobile,
            'tablet_count': tablet,
            'desktop_count': desktop,
        }

    def _get_recent_orders(self, store, limit=10):
        """Get recent orders for dashboard."""
        orders = Order.objects.filter(store=store).select_related(
            'user'
        ).order_by('-created_at')[:limit]
        
        return [{
            'id': o.id,
            'order_number': o.order_number,
            'customer_name': o.user.full_name if o.user else 'Anonymous',
            'total_price': float(o.total_price),
            'order_status': o.order_status,
            'created_at': o.created_at,
        } for o in orders]

    def _get_top_products(self, store, start_date, limit=5):
        """Get top selling products."""
        items = OrderItem.objects.filter(
            order__store=store,
            order__created_at__date__gte=start_date,
            order__order_status__in=['paid', 'processed', 'shipped', 'completed']
        ).values('product_id', 'product_name').annotate(
            total_sold=Sum('qty'),
            total_revenue=Sum('subtotal')
        ).order_by('-total_sold')[:limit]
        
        return [{
            'product_id': i['product_id'],
            'product_name': i['product_name'],
            'total_sold': i['total_sold'],
            'total_revenue': float(i['total_revenue']),
        } for i in items]


class SalesAnalyticsView(generics.ListAPIView):
    """Get detailed sales analytics."""
    serializer_class = SalesAnalyticsSerializer
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return SalesAnalytics.objects.none()

        store = self.request.user.store
        return SalesAnalytics.objects.filter(store=store).order_by('-date')


@extend_schema(exclude=True)
class SalesTrendDataView(views.APIView):
    """Get sales trend data for charts."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        store = request.user.store
        period = request.query_params.get('period', '30')
        days = int(period)
        
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)
        
        orders = Order.objects.filter(
            store=store,
            created_at__date__gte=start_date,
            order_status__in=['paid', 'completed']
        ).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            total=Sum('total_price'),
            count=Count('id')
        ).order_by('day')
        
        # Build chart data — use dict indexing instead of O(n*m) linear scan
        indexed = {}
        for o in orders:
            if o['day']:
                indexed[o['day'].date()] = o
        
        labels = []
        sales = []
        order_counts = []
        
        current = start_date
        while current <= end_date:
            labels.append(current.strftime('%d %b'))
            found = indexed.get(current)
            if found:
                sales.append(float(found['total']))
                order_counts.append(found['count'])
            else:
                sales.append(0)
                order_counts.append(0)
            current += timedelta(days=1)
        
        return Response({
            'labels': labels,
            'daily_sales': sales,
            'daily_orders': order_counts,
        })


@extend_schema(exclude=True)
class DeviceAnalyticsView(views.APIView):
    """Get device usage analytics."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        store = request.user.store
        period = request.query_params.get('period', '30')
        days = int(period)
        
        start_date = timezone.now().date() - timedelta(days=days)
        
        activities = UserActivity.objects.filter(
            store=store,
            created_at__date__gte=start_date
        )
        
        mobile = activities.filter(device_type='mobile').count()
        tablet = activities.filter(device_type='tablet').count()
        desktop = activities.filter(device_type='desktop').count()
        
        # Browser breakdown
        browsers = activities.values('device_type').annotate(
            count=Count('id')
        )
        
        return Response({
            'mobile': mobile,
            'tablet': tablet,
            'desktop': desktop,
            'total': mobile + tablet + desktop,
            'breakdown': [
                {'device': b['device_type'], 'count': b['count']}
                for b in browsers
            ],
        })


class UserActivityView(generics.ListAPIView):
    """Get recent user activities."""
    serializer_class = UserActivitySerializer
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get_queryset(self):
        store = self.request.user.store
        return UserActivity.objects.filter(store=store).select_related(
            'user'
        ).order_by('-created_at')[:50]


@extend_schema(exclude=True)
class DailyReportView(generics.ListAPIView):
    """Get daily reports."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get_queryset(self):
        store = self.request.user.store
        return DailyReport.objects.filter(store=store).order_by('-date')


@extend_schema(exclude=True)
class RealTimeAnalyticsView(views.APIView):
    """Get real-time analytics for WebSocket fallback."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        store = request.user.store
        today = timezone.now().date()
        
        today_orders = Order.objects.filter(
            store=store,
            created_at__date=today,
            order_status__in=['paid', 'processed', 'shipped', 'completed']
        )
        
        return Response({
            'today_sales': float(today_orders.aggregate(total=Sum('total_price'))['total'] or 0),
            'today_orders': today_orders.count(),
            'today_visitors': UserActivity.objects.filter(
                store=store,
                created_at__date=today,
                activity_type='page_view'
            ).count(),
            'pending_orders': Order.objects.filter(
                store=store, order_status='pending'
            ).count(),
        })

# =============================================================================
# AI BUSINESS INSIGHT
# =============================================================================


@extend_schema(exclude=True)
class AIBusinessInsightView(views.APIView):
    """Get AI-powered business insights for seller.

    Returns comprehensive analysis including:
    - Sales insights with growth trends
    - Product performance analysis
    - Customer behavior insights
    - Inventory health check
    - Actionable recommendations

    Flutter-ready JSON response.
    """
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        days = int(request.query_params.get('days', 30))
        service = AISellerInsightService(request.user.store)
        insights = service.get_comprehensive_insights(days)
        return Response(insights)


@extend_schema(exclude=True)
class AIQuickInsightView(views.APIView):
    """Get quick dashboard insights."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        service = AISellerInsightService(request.user.store)
        insights = service.get_quick_insights()
        return Response(insights)


@extend_schema(exclude=True)
class AIGrowthTipsView(views.APIView):
    """Get AI-generated growth tips."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        service = AISellerInsightService(request.user.store)
        tips = service.get_growth_tips()
        return Response(tips)


@extend_schema(exclude=True)
class AdminAIBusinessOverviewView(views.APIView):
    """Admin overview of all AI business insights."""
    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    def get(self, request):
        from stores.models import Store
        stores = Store.objects.filter(status='active')[:10]
        all_insights = []
        for store in stores:
            try:
                service = AISellerInsightService(store)
                quick = service.get_quick_insights()
                all_insights.append({
                    'store_id': store.id,
                    'store_name': store.store_name,
                    'quick_stats': quick.get('quick_stats', {}),
                    'alerts': quick.get('alerts', []),
                })
            except Exception:
                pass
        return Response({
            'stores_analyzed': len(all_insights),
            'insights': all_insights,
        })





@extend_schema(exclude=True)
class SellerReportView(views.APIView):
    """Live seller report built from existing order, product, review, and analytics records."""
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    COMPLETED_STATUSES = ['completed']

    def get(self, request):
        store = request.user.store
        start_date, end_date = self._date_range(request)
        days_count = max((end_date - start_date).days + 1, 1)
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=days_count - 1)

        current = self._period_metrics(store, start_date, end_date)
        previous = self._period_metrics(store, prev_start, prev_end)
        daily_sales = self._daily_sales(store, start_date, end_date)
        category_sales = self._category_sales(store, start_date, end_date)
        top_products = self._top_products(store, start_date, end_date)
        device_sales = self._device_breakdown(store, start_date, end_date)
        new_customer_days = self._new_customer_days(store, start_date, end_date)

        data = {
            'store': {
                'id': store.id,
                'store_name': store.store_name,
                'status': store.status,
            },
            'range': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat(),
                'days': days_count,
            },
            'metrics': current,
            'previous_metrics': previous,
            'trends': {
                'total_sales': self._pct_change(current['total_sales'], previous['total_sales']),
                'total_orders': self._pct_change(current['total_orders'], previous['total_orders']),
                'products_sold': self._pct_change(current['products_sold'], previous['products_sold']),
                'new_customers': self._pct_change(current['new_customers'], previous['new_customers']),
                'average_rating': round(current['average_rating'] - previous['average_rating'], 2),
            },
            'sales_chart': daily_sales,
            'category_sales': category_sales,
            'device_sales': device_sales,
            'top_products': top_products,
            'daily_sales': [
                {'date': row['date'], 'sales': row['sales'], 'orders': row['orders']}
                for row in daily_sales['rows']
            ],
            'performance': self._performance_summary(
                daily_sales['rows'], top_products, category_sales, device_sales, new_customer_days
            ),
        }
        return Response(data)

    def _date_range(self, request):
        today = timezone.localdate()
        period = request.query_params.get('period', '30days')
        if period == 'today':
            return today, today
        if period == '7days':
            return today - timedelta(days=6), today
        if period == 'this_month':
            return today.replace(day=1), today
        if period == 'this_year':
            return today.replace(month=1, day=1), today
        if period == 'custom':
            start_raw = request.query_params.get('start')
            end_raw = request.query_params.get('end')
            try:
                start = datetime.strptime(start_raw, '%Y-%m-%d').date()
                end = datetime.strptime(end_raw, '%Y-%m-%d').date()
                if start <= end:
                    return start, end
            except (TypeError, ValueError):
                pass
        return today - timedelta(days=29), today

    def _orders(self, store, start_date, end_date, completed_only=False):
        qs = Order.objects.filter(
            store=store,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        if completed_only:
            qs = qs.filter(order_status__in=self.COMPLETED_STATUSES)
        return qs

    def _period_metrics(self, store, start_date, end_date):
        completed = self._orders(store, start_date, end_date, completed_only=True)
        all_orders = self._orders(store, start_date, end_date)
        total_sales = completed.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
        products_sold = OrderItem.objects.filter(order__in=completed).aggregate(total=Sum('qty'))['total'] or 0
        new_customers = self._new_customers_count(store, start_date, end_date)
        average_rating = Review.objects.filter(
            product__store=store,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        ).aggregate(avg=Avg('rating'))['avg'] or 0
        return {
            'total_sales': float(total_sales),
            'total_orders': all_orders.count(),
            'products_sold': products_sold,
            'new_customers': new_customers,
            'average_rating': round(float(average_rating), 2),
        }

    def _new_customers_count(self, store, start_date, end_date):
        first_orders = Order.objects.filter(store=store, user__isnull=False).values('user').annotate(first=Min('created_at'))
        return sum(1 for row in first_orders if row['first'] and start_date <= row['first'].date() <= end_date)

    def _new_customer_days(self, store, start_date, end_date):
        first_orders = Order.objects.filter(store=store, user__isnull=False).values('user').annotate(first=Min('created_at'))
        counts = {}
        for row in first_orders:
            first = row['first']
            if first and start_date <= first.date() <= end_date:
                key = first.date().isoformat()
                counts[key] = counts.get(key, 0) + 1
        return counts

    def _daily_sales(self, store, start_date, end_date):
        rows = self._orders(store, start_date, end_date, completed_only=True).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            sales=Sum('total_price'),
            orders=Count('id'),
        ).order_by('day')
        indexed = {row['day'].date(): row for row in rows if row['day']}
        labels, sales, orders, table_rows = [], [], [], []
        current = start_date
        while current <= end_date:
            found = indexed.get(current)
            value = float(found['sales'] or 0) if found else 0
            count = found['orders'] if found else 0
            labels.append(current.strftime('%d %b'))
            sales.append(value)
            orders.append(count)
            table_rows.append({'date': current.isoformat(), 'label': current.strftime('%d %b %Y'), 'sales': value, 'orders': count})
            current += timedelta(days=1)
        return {'labels': labels, 'daily_sales': sales, 'daily_orders': orders, 'rows': table_rows}

    def _category_sales(self, store, start_date, end_date):
        rows = OrderItem.objects.filter(
            order__store=store,
            order__created_at__date__gte=start_date,
            order__created_at__date__lte=end_date,
            order__order_status__in=self.COMPLETED_STATUSES,
        ).values('product__category__category_name').annotate(
            revenue=Sum('subtotal'),
            quantity=Sum('qty'),
        ).order_by('-revenue')
        total = sum((row['revenue'] or Decimal('0')) for row in rows)
        return [{
            'category': row['product__category__category_name'] or 'Tanpa Kategori',
            'revenue': float(row['revenue'] or 0),
            'quantity': row['quantity'] or 0,
            'percentage': round(float((row['revenue'] or 0) / total * 100), 1) if total else 0,
        } for row in rows]

    def _top_products(self, store, start_date, end_date):
        rows = OrderItem.objects.filter(
            order__store=store,
            order__created_at__date__gte=start_date,
            order__created_at__date__lte=end_date,
            order__order_status__in=self.COMPLETED_STATUSES,
        ).values('product_id', 'product_name', 'product_photo').annotate(
            total_sold=Sum('qty'),
            total_revenue=Sum('subtotal'),
        ).order_by('-total_sold')[:5]
        return [{
            'product_id': row['product_id'],
            'product_name': row['product_name'] or 'Produk',
            'product_photo': row['product_photo'] or '',
            'total_sold': row['total_sold'] or 0,
            'total_revenue': float(row['total_revenue'] or 0),
        } for row in rows]

    def _device_breakdown(self, store, start_date, end_date):
        device_rows = DeviceAnalytics.objects.filter(
            store=store,
            date__gte=start_date,
            date__lte=end_date,
        ).values('device_type').annotate(count=Sum('visitors_count')).order_by('-count')

        if not device_rows:
            device_rows = UserActivity.objects.filter(
                store=store,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date,
            ).exclude(device_type__isnull=True).exclude(device_type='').values('device_type').annotate(count=Count('id')).order_by('-count')

        total = sum(row['count'] or 0 for row in device_rows)
        return [{
            'device': row['device_type'],
            'count': row['count'] or 0,
            'percentage': round(((row['count'] or 0) / total) * 100, 1) if total else 0,
        } for row in device_rows]

    def _performance_summary(self, daily_rows, top_products, category_sales, device_sales, new_customer_days):
        best_day = max(daily_rows, key=lambda row: row['sales'], default=None)
        top_category = category_sales[0] if category_sales else None
        top_device = device_sales[0] if device_sales else None
        best_customer_day = max(new_customer_days.items(), key=lambda item: item[1], default=None)
        return {
            'best_day': {
                'label': best_day['label'] if best_day else '-',
                'value': best_day['sales'] if best_day else 0,
                'orders': best_day['orders'] if best_day else 0,
            },
            'top_product': top_products[0] if top_products else None,
            'top_category': top_category,
            'top_device': top_device,
            'new_customers_peak': {
                'date': best_customer_day[0] if best_customer_day else None,
                'count': best_customer_day[1] if best_customer_day else 0,
            },
        }

    def _pct_change(self, current, previous):
        current = float(current or 0)
        previous = float(previous or 0)
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round(((current - previous) / previous) * 100, 1)


# =============================================================================
# REPORT EXPORT (PDF / CSV / XLSX)
# =============================================================================

import csv as csv_lib
from io import BytesIO
def _format_rupiah(val):
    """Format a number to Indonesian Rupiah string."""
    return f'Rp {float(val or 0):,.0f}'


def _write_csv_metric_row(writer, label, current_val, previous_val, pct_fn):
    """Write a comparison metric row with numeric values, trend, and pct_change."""
    pct = pct_fn(current_val, previous_val)
    trend = 'Naik' if current_val >= previous_val else 'Turun'
    writer.writerow([label, current_val, previous_val, pct, trend])


@extend_schema(exclude=True)
class ReportExportView(views.APIView):
    """
    Server-side report export endpoint.
    Usage: GET /api/analytics/export/?format=pdf&period=30days
    Supported formats: pdf, csv, xlsx
    """
    permission_classes = (permissions.IsAuthenticated, IsSeller)

    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        if fmt not in ('pdf', 'csv', 'xlsx'):
            return Response({'error': 'Format tidak didukung. Gunakan: pdf, csv, xlsx'}, status=400)

        # Build report data using SellerReportView logic
        report_view = SellerReportView()
        start_date, end_date = report_view._date_range(request)
        store = request.user.store
        days_count = max((end_date - start_date).days + 1, 1)
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=days_count - 1)

        current = report_view._period_metrics(store, start_date, end_date)
        previous = report_view._period_metrics(store, prev_start, prev_end)
        daily_sales = report_view._daily_sales(store, start_date, end_date)
        category_sales = report_view._category_sales(store, start_date, end_date)
        top_products = report_view._top_products(store, start_date, end_date)
        device_sales = report_view._device_breakdown(store, start_date, end_date)

        store_name = store.store_name
        period_label = f'{start_date.isoformat()} s.d. {end_date.isoformat()}'

        pct_fn = report_view._pct_change
        if fmt == 'csv':
            return self._generate_csv(store_name, period_label, current, previous, daily_sales, category_sales, top_products, pct_fn)
        elif fmt == 'xlsx':
            return self._generate_xlsx(store_name, period_label, current, previous, daily_sales, category_sales, top_products, pct_fn)
        else:
            return self._generate_pdf(store_name, period_label, current, previous, daily_sales, category_sales, top_products, device_sales, pct_fn)

    def _generate_csv(self, store_name, period_label, current, previous, daily_sales, category_sales, top_products, pct_fn):
        """Generate UTF-8 BOM CSV report for Microsoft Excel compatibility.

        - UTF-8 BOM (\xEF\xBB\xBF) signals Excel to open as UTF-8, preserving
          Indonesian characters (Kategori, Terjual, Penjualan, etc.)
        - Clean tabular sections with readable column headers
        - Raw numeric values (not stringified Rupiah) so Excel can sort/filter
        - No section markers (===) mixed into data rows
        """
        # Build the CSV content in-memory with utf-8-sig encoding (auto-prepends BOM)
        from io import StringIO

        buf = StringIO()
        writer = csv_lib.writer(buf)

        # ── Metadata header (2 rows, non-tabular, kept minimal) ──
        writer.writerow([f'Warungio — Laporan Penjualan: {store_name}'])
        writer.writerow([f'Periode: {period_label}'])

        # ── Section 1: Perbandingan Periode ──
        writer.writerow([])  # blank spacer
        writer.writerow(['Perbandingan Periode'])
        writer.writerow([
            'Metrik',
            'Periode Ini (Rp)',
            'Periode Sebelumnya (Rp)',
            'Perubahan (%)',
            'Trend',
        ])
        _write_csv_metric_row(writer, 'Total Penjualan', current['total_sales'],
                               previous['total_sales'], pct_fn)
        _write_csv_metric_row(writer, 'Total Pesanan', current['total_orders'],
                               previous['total_orders'], pct_fn)
        _write_csv_metric_row(writer, 'Produk Terjual', current['products_sold'],
                               previous['products_sold'], pct_fn)
        _write_csv_metric_row(writer, 'Pelanggan Baru', current['new_customers'],
                               previous['new_customers'], pct_fn)
        # Rating (not a pct change, use absolute diff)
        rating_diff = round(current['average_rating'] - previous['average_rating'], 2)
        trend_rating = 'Naik' if rating_diff >= 0 else 'Turun'
        writer.writerow([
            'Rata-rata Rating',
            current['average_rating'],
            previous['average_rating'],
            rating_diff,
            trend_rating,
        ])

        # ── Section 2: Penjualan Harian ──
        writer.writerow([])
        writer.writerow(['Penjualan Harian'])
        writer.writerow(['Tanggal', 'Penjualan (Rp)', 'Pesanan'])
        daily_rows = daily_sales.get('rows', [])
        for row in daily_rows:
            writer.writerow([row['label'], row['sales'], row['orders']])
        # Totals row
        if daily_rows:
            total_sales = sum(r['sales'] for r in daily_rows)
            total_orders = sum(r['orders'] for r in daily_rows)
            writer.writerow(['TOTAL', total_sales, total_orders])

        # ── Section 3: Penjualan per Kategori ──
        writer.writerow([])
        writer.writerow(['Penjualan per Kategori'])
        writer.writerow(['Kategori', 'Penjualan (Rp)', 'Jumlah Terjual', 'Persentase (%)'])
        for cat in category_sales:
            writer.writerow([cat['category'], cat['revenue'], cat['quantity'], cat['percentage']])
        # Totals row
        if category_sales:
            cat_total_rev = sum(c['revenue'] for c in category_sales)
            cat_total_qty = sum(c['quantity'] for c in category_sales)
            writer.writerow(['TOTAL', cat_total_rev, cat_total_qty, 100.0])

        # ── Section 4: Produk Terlaris ──
        writer.writerow([])
        writer.writerow(['Produk Terlaris'])
        writer.writerow(['Peringkat', 'Produk', 'Terjual', 'Pendapatan (Rp)'])
        for idx, p in enumerate(top_products, 1):
            writer.writerow([f'#{idx}', p['product_name'], p['total_sold'], p['total_revenue']])
        # Totals row
        if top_products:
            prod_total_sold = sum(p['total_sold'] for p in top_products)
            prod_total_rev = sum(p['total_revenue'] for p in top_products)
            writer.writerow(['', 'TOTAL', prod_total_sold, prod_total_rev])

        # ── Encode with UTF-8 BOM ──
        raw = buf.getvalue()
        buf.close()

        response = HttpResponse(
            raw.encode('utf-8-sig'),  # utf-8-sig prepends \xEF\xBB\xBF BOM
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="laporan-{store_name.lower().replace(" ", "-")}.csv"'
        )
        # Explicit charset in Content-Type + BOM = Excel opens correctly
        response['Content-Type'] = 'text/csv; charset=utf-8'
        return response

    def _generate_xlsx(self, store_name, period_label, current, previous, daily_sales, category_sales, top_products, pct_fn):
        """Generate a professional Excel (.xlsx) report with multiple worksheets,
        Warungio branding, currency formatting, auto-sized columns, and totals."""
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side
        )
        from openpyxl.utils import get_column_letter
        from datetime import datetime as dt

        wb = Workbook()

        # ── Colour palette ──
        GREEN_PRIMARY = '16A34A'
        GREEN_DARK = '15803D'
        GREEN_LIGHT = 'DCFCE7'
        GREEN_BG = 'F0FDF4'
        WHITE = 'FFFFFF'
        GREY_DARK = '1E293B'
        GREY_MED = '64748B'
        GREY_LIGHT = 'E2E8F0'
        GREY_BG = 'F8FAFC'
        AMBER_LIGHT = 'FEF3C7'
        AMBER_BG = 'FFFBEB'

        # ── Reusable styles ──
        green_fill = PatternFill(start_color=GREEN_PRIMARY, end_color=GREEN_PRIMARY, fill_type='solid')
        green_light_fill = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type='solid')
        green_bg_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type='solid')
        grey_bg_fill = PatternFill(start_color=GREY_BG, end_color=GREY_BG, fill_type='solid')
        amber_light_fill = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type='solid')
        white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color=GREY_LIGHT),
            right=Side(style='thin', color=GREY_LIGHT),
            top=Side(style='thin', color=GREY_LIGHT),
            bottom=Side(style='thin', color=GREY_LIGHT),
        )
        green_bottom_border = Border(
            left=Side(style='thin', color=GREY_LIGHT),
            right=Side(style='thin', color=GREY_LIGHT),
            top=Side(style='thin', color=GREY_LIGHT),
            bottom=Side(style='medium', color=GREEN_PRIMARY),
        )
        header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_font = Font(name='Calibri', size=16, bold=True, color=GREEN_DARK)
        subtitle_font = Font(name='Calibri', size=10, italic=True, color=GREY_MED)
        section_font = Font(name='Calibri', size=12, bold=True, color=GREEN_DARK)
        total_font = Font(name='Calibri', size=11, bold=True, color=GREY_DARK)
        label_font = Font(name='Calibri', size=10, color=GREY_DARK)
        label_bold_font = Font(name='Calibri', size=10, bold=True, color=GREY_DARK)
        meta_label_font = Font(name='Calibri', size=9, color=GREY_MED)
        meta_value_font = Font(name='Calibri', size=9, bold=True, color=GREY_DARK)
        data_align_left = Alignment(horizontal='left', vertical='center')
        data_align_center = Alignment(horizontal='center', vertical='center')
        data_align_right = Alignment(horizontal='right', vertical='center')

        currency_fmt = '#,##0'
        pct_fmt = '0.0%'
        date_fmt = 'DD MMM YYYY'

        # ── Helper: auto-size columns ──
        def auto_size(ws, min_width=10, max_width=45):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    val = cell.value
                    if val is None:
                        continue
                    # Approximate: count chars, consider CJK chars wider
                    s = str(val)
                    length = 0
                    for ch in s:
                        if ord(ch) > 0x2E80:
                            length += 2  # wide char
                        else:
                            length += 1
                    if cell.font and cell.font.bold:
                        length += 1
                    max_len = max(max_len, length)
                adjusted = min(max(max_len + 3, min_width), max_width)
                ws.column_dimensions[col_letter].width = adjusted

        # ── Helper: write a styled data row with zebra ──
        def write_data_row(ws, row_num, values, formats=None,
                           zebra=False, is_total=False, is_header=False):
            """Write a data row with border, optional zebra, and per-column formats.
            formats is a list of dicts with keys: 'align', 'font', 'number_format'.
            """
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)

                if is_header:
                    cell.font = header_font
                    cell.fill = green_fill
                    cell.alignment = header_align
                elif is_total:
                    cell.font = total_font
                    cell.fill = green_light_fill
                    cell.alignment = data_align_center
                else:
                    cell.font = label_font
                    cell.alignment = data_align_center
                    if zebra and (row_num % 2 == 0):
                        cell.fill = grey_bg_fill

                cell.border = thin_border

                # Apply per-column formatting
                if formats and col_idx - 1 < len(formats):
                    fmt = formats[col_idx - 1]
                    if 'align' in fmt:
                        cell.alignment = fmt['align']
                    if 'number_format' in fmt:
                        cell.number_format = fmt['number_format']
                    if is_header or is_total:
                        pass  # keep header/total fonts
                    elif 'font' in fmt:
                        cell.font = fmt['font']

            return row_num + 1

        # ════════════════════════════════════════════════════════════════
        # SHEET 1 — RINGKASAN (Summary)
        # ════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.sheet_properties.tabColor = GREEN_PRIMARY

        now_str = dt.now().strftime('%d/%m/%Y %H:%M')

        # -- Title row --
        ws1.merge_cells('A1:F1')
        title_cell = ws1.cell(row=1, column=1, value=f'Warungio — Laporan Penjualan')
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws1.row_dimensions[1].height = 30

        # -- Branding separator --
        ws1.merge_cells('A2:F2')
        sep_cell = ws1.cell(row=2, column=1, value='')
        sep_cell.fill = green_fill
        ws1.row_dimensions[2].height = 4

        # -- Metadata block --
        r = 4
        metadata = [
            ('Toko', store_name),
            ('Periode', period_label),
            ('Dicetak', f'{now_str} WIB'),
            ('Status', 'Aktif'),
        ]
        for label, value in metadata:
            cell_lbl = ws1.cell(row=r, column=1, value=label)
            cell_lbl.font = meta_label_font
            cell_val = ws1.cell(row=r, column=2, value=value)
            cell_val.font = meta_value_font
            ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            r += 1

        r += 1  # spacer

        # -- Section: Perbandingan Periode --
        section_row = r
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        sec_cell = ws1.cell(row=r, column=1, value='PERBANDINGAN PERIODE')
        sec_cell.font = section_font
        r += 1

        # Table headers
        headers = ['Metrik', 'Periode Ini', 'Periode Sebelumnya', 'Perubahan', 'Trend']
        header_formats = [
            {'align': data_align_left},
            {'align': data_align_right},
            {'align': data_align_right},
            {'align': data_align_center},
            {'align': data_align_center},
        ]
        r = write_data_row(ws1, r, headers, formats=header_formats, is_header=True)

        # Data rows
        rows_data = [
            ('Total Penjualan', current['total_sales'], previous['total_sales'],
             pct_fn(current['total_sales'], previous['total_sales']),
             'Naik' if current['total_sales'] >= previous['total_sales'] else 'Turun',
             [{'align': data_align_left, 'number_format': currency_fmt},
              {'align': data_align_right, 'number_format': currency_fmt},
              {'align': data_align_right},
              {'align': data_align_center},
              {'align': data_align_center}]),
            ('Total Pesanan', current['total_orders'], previous['total_orders'],
             pct_fn(current['total_orders'], previous['total_orders']),
             'Naik' if current['total_orders'] >= previous['total_orders'] else 'Turun',
             [{'align': data_align_left},
              {'align': data_align_right},
              {'align': data_align_right},
              {'align': data_align_center},
              {'align': data_align_center}]),
            ('Produk Terjual', current['products_sold'], previous['products_sold'],
             pct_fn(current['products_sold'], previous['products_sold']),
             'Naik' if current['products_sold'] >= previous['products_sold'] else 'Turun',
             [{'align': data_align_left},
              {'align': data_align_right},
              {'align': data_align_right},
              {'align': data_align_center},
              {'align': data_align_center}]),
            ('Pelanggan Baru', current['new_customers'], previous['new_customers'],
             pct_fn(current['new_customers'], previous['new_customers']),
             'Naik' if current['new_customers'] >= previous['new_customers'] else 'Turun',
             [{'align': data_align_left},
              {'align': data_align_right},
              {'align': data_align_right},
              {'align': data_align_center},
              {'align': data_align_center}]),
            ('Rata-rata Rating', current['average_rating'], previous['average_rating'],
             round(current['average_rating'] - previous['average_rating'], 2),
             'Naik' if current['average_rating'] >= previous['average_rating'] else 'Turun',
             [{'align': data_align_left},
              {'align': data_align_right, 'number_format': '0.00'},
              {'align': data_align_right, 'number_format': '0.00'},
              {'align': data_align_center, 'number_format': '0.00'},
              {'align': data_align_center}]),
        ]
        for rd in rows_data:
            label = rd[0]
            v1 = rd[1]
            v2 = rd[2]
            pct = float(rd[3]) / 100  # store as decimal for pct format
            trend = rd[4]
            fmts = rd[5]
            values = [label, v1, v2, pct, trend]
            write_data_row(ws1, r, values, formats=fmts, zebra=True)
            r += 1

        # -- Totals row for comparison --
        total_current_sales = current['total_sales']
        total_previous_sales = previous['total_sales']
        total_current_pct = pct_fn(total_current_sales, total_previous_sales)
        total_fmts = [
            {'align': data_align_left, 'font': total_font},
            {'align': data_align_right, 'number_format': currency_fmt},
            {'align': data_align_right, 'number_format': currency_fmt},
            {'align': data_align_center},
            {'align': data_align_center},
        ]
        write_data_row(ws1, r, [
            'TOTAL', total_current_sales, total_previous_sales,
            float(total_current_pct) / 100,
            '--'
        ], formats=total_fmts, is_total=True)

        # Freeze panes at row 4 (metadata visible, table scrollable)
        ws1.freeze_panes = 'A4'

        # Auto-filter on comparison table
        last_data_row = r
        ws1.auto_filter.ref = f'A{section_row + 1}:E{last_data_row}'

        # ════════════════════════════════════════════════════════════════
        # SHEET 2 — PENJUALAN HARIAN (Daily Sales)
        # ════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Penjualan Harian')
        ws2.sheet_properties.tabColor = GREEN_PRIMARY

        r2 = 1
        # Title
        ws2.merge_cells('A1:C1')
        t2 = ws2.cell(row=1, column=1, value=f'Penjualan Harian — {store_name}')
        t2.font = title_font
        ws2.row_dimensions[1].height = 28

        ws2.merge_cells('A2:C2')
        sub2 = ws2.cell(row=2, column=1, value=f'Periode: {period_label}  |  Dicetak: {now_str} WIB')
        sub2.font = subtitle_font

        # Header
        r2 = 4
        daily_headers = ['Tanggal', 'Penjualan (Rp)', 'Pesanan']
        daily_fmts = [{'align': data_align_left}, {'align': data_align_right}, {'align': data_align_center}]
        r2 = write_data_row(ws2, r2, daily_headers, formats=daily_fmts, is_header=True)

        daily_rows = daily_sales.get('rows', [])
        total_sales = 0
        total_orders = 0
        for row in daily_rows:
            vals = [row['label'], row['sales'], row['orders']]
            fmts = [
                {'align': data_align_left, 'number_format': date_fmt},
                {'align': data_align_right, 'number_format': currency_fmt},
                {'align': data_align_center},
            ]
            write_data_row(ws2, r2, vals, formats=fmts, zebra=True)
            total_sales += row['sales']
            total_orders += row['orders']
            r2 += 1

        # Totals row
        total_daily_fmts = [
            {'align': data_align_left, 'font': total_font},
            {'align': data_align_right, 'number_format': currency_fmt},
            {'align': data_align_center},
        ]
        write_data_row(ws2, r2, ['TOTAL', total_sales, total_orders],
                       formats=total_daily_fmts, is_total=True)

        ws2.freeze_panes = 'A4'
        ws2.auto_filter.ref = f'A4:C{r2}'

        # ════════════════════════════════════════════════════════════════
        # SHEET 3 — KATEGORI (Category Sales)
        # ════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Kategori')
        ws3.sheet_properties.tabColor = GREEN_PRIMARY

        r3 = 1
        ws3.merge_cells('A1:D1')
        t3 = ws3.cell(row=1, column=1, value='Penjualan per Kategori')
        t3.font = title_font
        ws3.row_dimensions[1].height = 28

        ws3.merge_cells('A2:D2')
        sub3 = ws3.cell(row=2, column=1, value=f'{store_name} — {period_label}')
        sub3.font = subtitle_font

        r3 = 4
        cat_headers = ['Kategori', 'Penjualan (Rp)', 'Jumlah Terjual', 'Persentase']
        cat_fmts_header = [
            {'align': data_align_left}, {'align': data_align_right},
            {'align': data_align_center}, {'align': data_align_center},
        ]
        r3 = write_data_row(ws3, r3, cat_headers, formats=cat_fmts_header, is_header=True)

        cat_total_rev = 0
        cat_total_qty = 0
        for cat in category_sales:
            pct_decimal = cat['percentage'] / 100
            vals = [cat['category'], cat['revenue'], cat['quantity'], pct_decimal]
            fmts = [
                {'align': data_align_left},
                {'align': data_align_right, 'number_format': currency_fmt},
                {'align': data_align_center},
                {'align': data_align_center, 'number_format': '0.0%'},
            ]
            write_data_row(ws3, r3, vals, formats=fmts, zebra=True)
            cat_total_rev += cat['revenue']
            cat_total_qty += cat['quantity']
            r3 += 1

        # Totals row
        cat_total_fmts = [
            {'align': data_align_left, 'font': total_font},
            {'align': data_align_right, 'number_format': currency_fmt},
            {'align': data_align_center},
            {'align': data_align_center, 'number_format': '0.0%'},
        ]
        write_data_row(ws3, r3, ['TOTAL', cat_total_rev, cat_total_qty, 1.0],
                       formats=cat_total_fmts, is_total=True)

        ws3.freeze_panes = 'A4'
        ws3.auto_filter.ref = f'A4:D{r3}'

        # ════════════════════════════════════════════════════════════════
        # SHEET 4 — PRODUK TERLARIS (Top Products)
        # ════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('Produk Terlaris')
        ws4.sheet_properties.tabColor = GREEN_PRIMARY

        r4 = 1
        ws4.merge_cells('A1:D1')
        t4 = ws4.cell(row=1, column=1, value='Produk Terlaris')
        t4.font = title_font
        ws4.row_dimensions[1].height = 28

        ws4.merge_cells('A2:D2')
        sub4 = ws4.cell(row=2, column=1, value=f'{store_name} — {period_label}')
        sub4.font = subtitle_font

        r4 = 4
        prod_headers = ['Peringkat', 'Produk', 'Terjual', 'Pendapatan (Rp)']
        prod_fmts_header = [
            {'align': data_align_center}, {'align': data_align_left},
            {'align': data_align_center}, {'align': data_align_right},
        ]
        r4 = write_data_row(ws4, r4, prod_headers, formats=prod_fmts_header, is_header=True)

        prod_total_sold = 0
        prod_total_rev = 0
        for idx, p in enumerate(top_products, 1):
            vals = [f'#{idx}', p['product_name'], p['total_sold'], p['total_revenue']]
            fmts = [
                {'align': data_align_center},
                {'align': data_align_left},
                {'align': data_align_center},
                {'align': data_align_right, 'number_format': currency_fmt},
            ]
            write_data_row(ws4, r4, vals, formats=fmts, zebra=True)
            prod_total_sold += p['total_sold']
            prod_total_rev += p['total_revenue']
            r4 += 1

        # Totals row
        prod_total_fmts = [
            {'align': data_align_center, 'font': total_font},
            {'align': data_align_left, 'font': total_font},
            {'align': data_align_center},
            {'align': data_align_right, 'number_format': currency_fmt},
        ]
        write_data_row(ws4, r4, ['', 'TOTAL', prod_total_sold, prod_total_rev],
                       formats=prod_total_fmts, is_total=True)

        ws4.freeze_panes = 'A4'
        ws4.auto_filter.ref = f'A4:D{r4}'

        # ── Auto-size columns on all sheets ──
        for ws in [ws1, ws2, ws3, ws4]:
            auto_size(ws, min_width=10, max_width=50)

        # ── Build response ──
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = store_name.lower().replace(' ', '-')[:50]
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="laporan-{safe_name}.xlsx"'
        )
        return response

    def _generate_pdf(self, store_name, period_label, current, previous, daily_sales, category_sales, top_products, device_sales, pct_fn):
        """Generate a production-ready A4 PDF report with logo, header, footer, page numbers."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import (
            Paragraph, Spacer, Table, TableStyle,
            Image, PageBreak,
        )
        from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate, Frame
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from django.conf import settings as django_settings
        from datetime import datetime as dt
        import os

        # ── Colours ──
        GREEN_PRIMARY = colors.HexColor('#16A34A')
        GREEN_DARK = colors.HexColor('#15803D')
        GREEN_LIGHT = colors.HexColor('#F0FDF4')
        GREEN_MUTED = colors.HexColor('#DCFCE7')
        WHITE = colors.white
        BLACK = colors.black
        GREY_DARK = colors.HexColor('#1E293B')
        GREY_MED = colors.HexColor('#64748B')
        GREY_LIGHT = colors.HexColor('#E2E8F0')
        GREY_BG = colors.HexColor('#F8FAFC')
        CARD_BG = colors.HexColor('#F0FDF4')
        UP_GREEN = colors.HexColor('#16A34A')
        DOWN_RED = colors.HexColor('#DC2626')
        NEUTRAL_GREY = colors.HexColor('#64748B')

        # ── Page size & margins ──
        PAGE_W, PAGE_H = A4
        MARGIN_L = 2 * cm
        MARGIN_R = 2 * cm
        MARGIN_T = 1.8 * cm
        MARGIN_B = 2.2 * cm
        HEADER_H = 3.2 * cm
        FOOTER_H = 1.8 * cm
        usable_w = PAGE_W - MARGIN_L - MARGIN_R
        usable_h = PAGE_H - MARGIN_T - MARGIN_B - HEADER_H - FOOTER_H

        # ── Logo path ──
        logo_candidates = [
            os.path.join(django_settings.BASE_DIR, 'assets', 'images', 'Warungio L.png'),
            os.path.join(django_settings.BASE_DIR, 'django_backend', 'static', 'images', 'Warungio L.png'),
            os.path.join(django_settings.BASE_DIR, 'django_backend', 'static', 'images', 'Warungio-L.png'),
        ]
        logo_path = None
        for candidate in logo_candidates:
            if os.path.exists(candidate):
                logo_path = candidate
                break

        buf = BytesIO()

        # ── Styles ──
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('PDF_Title', parent=styles['Title'],
            fontSize=20, leading=24, spaceAfter=2, textColor=GREEN_PRIMARY,
            fontName='Helvetica-Bold')
        s_subtitle = ParagraphStyle('PDF_Subtitle', parent=styles['Normal'],
            fontSize=10, leading=13, textColor=GREY_MED, spaceAfter=4)
        s_meta = ParagraphStyle('PDF_Meta', parent=styles['Normal'],
            fontSize=8, leading=10, textColor=GREY_MED, spaceAfter=2)
        s_section = ParagraphStyle('PDF_Section', parent=styles['Heading2'],
            fontSize=13, leading=16, textColor=GREEN_PRIMARY,
            spaceBefore=14, spaceAfter=6, fontName='Helvetica-Bold')
        s_card_val = ParagraphStyle('CardVal', parent=styles['Normal'],
            fontSize=13, leading=16, textColor=GREY_DARK, fontName='Helvetica-Bold')
        s_card_lbl = ParagraphStyle('CardLbl', parent=styles['Normal'],
            fontSize=7, leading=9, textColor=GREY_MED)
        s_card_pct = ParagraphStyle('CardPct', parent=styles['Normal'],
            fontSize=8, leading=10, fontName='Helvetica')
        s_body = ParagraphStyle('PDF_Body', parent=styles['Normal'],
            fontSize=8, leading=11, textColor=GREY_DARK)
        s_total = ParagraphStyle('PDF_Total', parent=styles['Normal'],
            fontSize=9, leading=12, textColor=GREY_DARK, fontName='Helvetica-Bold')
        s_footer = ParagraphStyle('PDF_Footer', parent=styles['Normal'],
            fontSize=7, leading=10, textColor=colors.HexColor('#94A3B8'),
            alignment=TA_CENTER)
        s_header_right = ParagraphStyle('HeaderRight', parent=styles['Normal'],
            fontSize=8, leading=10, textColor=GREY_MED, alignment=TA_RIGHT)

        # ── Header / Footer canvas handler ──
        _page_num = [0]

        def header_footer(canvas_obj, doc_obj):
            _page_num[0] += 1
            canvas_obj.saveState()

            # ── Header ──
            # Green header bar
            canvas_obj.setFillColor(GREEN_PRIMARY)
            canvas_obj.roundRect(MARGIN_L, PAGE_H - MARGIN_T - HEADER_H + 0.3*cm,
                                 usable_w, HEADER_H - 0.6*cm, 6, fill=1, stroke=0)
            # Logo
            if logo_path and os.path.exists(logo_path):
                try:
                    img = Image(logo_path, width=2.8*cm, height=1.8*cm)
                    img.drawOn(canvas_obj, MARGIN_L + 0.6*cm,
                               PAGE_H - MARGIN_T - HEADER_H + 0.7*cm)
                except Exception:
                    pass
            # Title text on header
            canvas_obj.setFillColor(WHITE)
            canvas_obj.setFont('Helvetica-Bold', 16)
            canvas_obj.drawString(MARGIN_L + 3.6*cm,
                                  PAGE_H - MARGIN_T - HEADER_H + 1.6*cm,
                                  'LAPORAN PENJUALAN')
            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.drawString(MARGIN_L + 3.6*cm,
                                  PAGE_H - MARGIN_T - HEADER_H + 1.0*cm,
                                  f'{store_name}  |  {period_label}')
            # Top-right timestamp
            now_str = dt.now().strftime('%d/%m/%Y %H:%M')
            canvas_obj.setFont('Helvetica', 7)
            canvas_obj.drawRightString(PAGE_W - MARGIN_R - 0.4*cm,
                                       PAGE_H - MARGIN_T - HEADER_H + 1.3*cm,
                                       f'Dicetak: {now_str} WIB')

            # ── Footer ──
            canvas_obj.setStrokeColor(GREY_LIGHT)
            canvas_obj.setLineWidth(0.5)
            canvas_obj.line(MARGIN_L, MARGIN_B + FOOTER_H - 0.8*cm,
                            PAGE_W - MARGIN_R, MARGIN_B + FOOTER_H - 0.8*cm)
            canvas_obj.setFillColor(GREY_MED)
            canvas_obj.setFont('Helvetica', 7)
            canvas_obj.drawString(MARGIN_L, MARGIN_B + 0.2*cm,
                                  'Warungio Seller Dashboard — Laporan ini digenerate secara otomatis')
            # Page X of Y
            canvas_obj.drawRightString(PAGE_W - MARGIN_R,
                                       MARGIN_B + 0.2*cm,
                                       f'Halaman {_page_num[0]}')
            canvas_obj.restoreState()

        # ── Document ──
        frame = Frame(MARGIN_L, MARGIN_B + FOOTER_H,
                      usable_w, usable_h,
                      leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
                      id='normal')
        template = PageTemplate(id='main', frames=[frame],
                                onPage=header_footer, pagesize=A4)

        doc = BaseDocTemplate(
            buf, pagesize=A4,
            leftMargin=MARGIN_L, rightMargin=MARGIN_R,
            topMargin=MARGIN_T + HEADER_H,
            bottomMargin=MARGIN_B + FOOTER_H,
        )
        doc.addPageTemplates([template])

        elements = []

        # ═══════════════════════════════════════════════════════════════════
        # 1.  TITLE BLOCK
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph('RINGKASAN KINERJA', s_title))
        elements.append(Paragraph(store_name, s_subtitle))
        elements.append(Paragraph(f'Periode laporan: {period_label}', s_meta))
        elements.append(Paragraph(
            f'Dicetak: {dt.now().strftime("%d %B %Y pukul %H:%M")} WIB | '
            f'Seller: {store_name}',
            s_meta
        ))
        elements.append(Spacer(1, 4*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 2.  SUMMARY CARDS (2x3 grid of visual cards)
        # ═══════════════════════════════════════════════════════════════════
        card_w = usable_w / 3 - 6
        card_h = 2.2 * cm

        def make_card(icon, label, value_str, trend_str, trend_val, good_up=True):
            """Build a bordered card table cell."""
            if trend_val > 0:
                trend_color = UP_GREEN if good_up else DOWN_RED
                arrow = '\u25B2'  # up triangle
            elif trend_val < 0:
                trend_color = DOWN_RED if good_up else UP_GREEN
                arrow = '\u25BC'  # down triangle
            else:
                trend_color = NEUTRAL_GREY
                arrow = '\u25C6'  # diamond
            trend_text = f'{arrow} {abs(trend_val)}%'
            cell = Table([
                [Paragraph(icon, ParagraphStyle('Icon', fontSize=16, leading=18, textColor=GREEN_PRIMARY))],
                [Paragraph(value_str, s_card_val)],
                [Paragraph(label, s_card_lbl)],
                [Paragraph(
                    f'<font color="{trend_color.hexval()}">{trend_text}</font> '
                    f'<font color="{GREY_MED.hexval()}" size="6">dari periode sblm</font>',
                    ParagraphStyle('Trend', fontSize=7, leading=9, textColor=trend_color)
                )],
            ], colWidths=[card_w], rowHeights=[18, 18, 12, 14])
            cell.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('BACKGROUND', (0, 0), (-1, -1), WHITE),
                ('BOX', (0, 0), (-1, -1), 0.6, GREY_LIGHT),
            ]))
            return cell

        # Build 3 cards per row, 2 rows (using simple text symbols, not emoji,
        # because Helvetica doesn't support emoji glyphs in PDF)
        metrics_cards = [
            ('$', 'Total Penjualan', _format_rupiah(current['total_sales']),
             pct_fn(current['total_sales'], previous['total_sales']), True),
            ('#', 'Total Pesanan', str(current['total_orders']),
             pct_fn(current['total_orders'], previous['total_orders']), True),
            ('*', 'Produk Terjual', str(current['products_sold']),
             pct_fn(current['products_sold'], previous['products_sold']), True),
            ('@', 'Pelanggan Baru', str(current['new_customers']),
             pct_fn(current['new_customers'], previous['new_customers']), True),
            ('+', 'Rating', f'{current["average_rating"]}/5',
             round(current['average_rating'] - previous['average_rating'], 2) * 10, True),
            ('>>', 'Rata-rata Pesanan',
             _format_rupiah(current['total_sales'] / max(current['total_orders'], 1)),
             0, True),
        ]

        card_elements = []
        for i in range(0, 6, 3):
            row_data = []
            for j in range(3):
                icon, lbl, val, trend, good = metrics_cards[i + j]
                card = make_card(icon, lbl, val, trend, trend, good)
                row_data.append(card)
            card_table = Table([row_data], colWidths=[card_w, card_w, card_w])
            card_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            card_elements.append(card_table)

        cards_combined = Table([[c] for c in card_elements], colWidths=[usable_w])
        cards_combined.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(cards_combined)
        elements.append(Spacer(1, 5*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 3.  PERIOD COMPARISON TABLE
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph('PERBANDINGAN PERIODE', s_section))

        def make_table_style(has_total=False):
            cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), GREEN_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.4, GREY_LIGHT),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, GREEN_LIGHT]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            if has_total:
                cmds += [
                    ('BACKGROUND', (0, -1), (-1, -1), GREEN_MUTED),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, GREEN_PRIMARY),
                ]
            return TableStyle(cmds)

        comp_data = [
            ['Metrik', 'Periode Ini', 'Periode Sebelumnya', 'Perubahan', 'Trend'],
            ['Total Penjualan',
             _format_rupiah(current['total_sales']),
             _format_rupiah(previous['total_sales']),
             f"{pct_fn(current['total_sales'], previous['total_sales'])}%",
             '\u25B2 Naik' if current['total_sales'] >= previous['total_sales'] else '\u25BC Turun'],
            ['Total Pesanan',
             str(current['total_orders']),
             str(previous['total_orders']),
             f"{pct_fn(current['total_orders'], previous['total_orders'])}%",
             '\u25B2 Naik' if current['total_orders'] >= previous['total_orders'] else '\u25BC Turun'],
            ['Produk Terjual',
             str(current['products_sold']),
             str(previous['products_sold']),
             f"{pct_fn(current['products_sold'], previous['products_sold'])}%",
             '\u25B2 Naik' if current['products_sold'] >= previous['products_sold'] else '\u25BC Turun'],
            ['Pelanggan Baru',
             str(current['new_customers']),
             str(previous['new_customers']),
             f"{pct_fn(current['new_customers'], previous['new_customers'])}%",
             '\u25B2 Naik' if current['new_customers'] >= previous['new_customers'] else '\u25BC Turun'],
            ['Rata-rata Rating',
             str(current['average_rating']),
             str(previous['average_rating']),
             str(round(current['average_rating'] - previous['average_rating'], 2)),
             '\u25B2 Naik' if current['average_rating'] >= previous['average_rating'] else '\u25BC Turun'],
        ]
        t_comp = Table(comp_data, colWidths=[120, 90, 90, 70, 60])
        t_comp.setStyle(make_table_style())
        elements.append(t_comp)
        elements.append(Spacer(1, 5*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 4.  DAILY SALES TABLE (with totals row)
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph('PENJUALAN HARIAN', s_section))
        daily_rows = daily_sales.get('rows', [])

        # Paginate daily rows: max 25 per page
        page_size = 25
        daily_pages = [daily_rows[i:i + page_size] for i in range(0, len(daily_rows), page_size)]

        for page_idx, page_rows in enumerate(daily_pages):
            total_sales = sum(r['sales'] for r in page_rows)
            total_orders = sum(r['orders'] for r in page_rows)

            daily_data = [['Tanggal', 'Penjualan', 'Pesanan']]
            for row in page_rows:
                daily_data.append([
                    row['label'],
                    _format_rupiah(row['sales']),
                    str(row['orders']),
                ])
            # Totals row
            daily_data.append([
                Paragraph('<b>TOTAL</b>', ParagraphStyle('TotalLabel', fontSize=8, leading=10, fontName='Helvetica-Bold')),
                Paragraph(f'<b>{_format_rupiah(total_sales)}</b>', ParagraphStyle('TotalVal', fontSize=8, leading=10, fontName='Helvetica-Bold')),
                Paragraph(f'<b>{total_orders}</b>', ParagraphStyle('TotalOrd', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            ])

            if page_idx < len(daily_pages) - 1:
                # Sub-label for paginated pages
                label = f'(halaman {page_idx + 1} dari {len(daily_pages)})'
                elements.append(Paragraph(label, ParagraphStyle(
                    'PageSub', fontSize=7, leading=9, textColor=GREY_MED,
                    alignment=TA_RIGHT, spaceAfter=2)))

            t_daily = Table(daily_data, colWidths=[120, 120, 80])
            t_daily.setStyle(make_table_style(has_total=True))
            elements.append(t_daily)

            if page_idx < len(daily_pages) - 1:
                elements.append(PageBreak())

        elements.append(Spacer(1, 5*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 5.  CATEGORY SALES TABLE
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph('PENJUALAN PER KATEGORI', s_section))
        cat_total_rev = sum(c['revenue'] for c in category_sales)
        cat_total_qty = sum(c['quantity'] for c in category_sales)
        cat_data = [['Kategori', 'Penjualan', 'Jumlah', 'Persentase']]
        for cat in category_sales:
            cat_data.append([
                cat['category'],
                _format_rupiah(cat['revenue']),
                str(cat['quantity']),
                f"{cat['percentage']}%",
            ])
        cat_data.append([
            Paragraph('<b>TOTAL</b>', ParagraphStyle('CatTotalLbl', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            Paragraph(f'<b>{_format_rupiah(cat_total_rev)}</b>', ParagraphStyle('CatTotalRev', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            Paragraph(f'<b>{cat_total_qty}</b>', ParagraphStyle('CatTotalQty', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            Paragraph('<b>100%</b>', ParagraphStyle('CatTotalPct', fontSize=8, leading=10, fontName='Helvetica-Bold')),
        ])
        t_cat = Table(cat_data, colWidths=[110, 90, 60, 70])
        t_cat.setStyle(make_table_style(has_total=True))
        elements.append(t_cat)
        elements.append(Spacer(1, 5*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 6.  TOP PRODUCTS TABLE
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph('PRODUK TERLARIS', s_section))
        prod_total_sold = sum(p['total_sold'] for p in top_products)
        prod_total_rev = sum(p['total_revenue'] for p in top_products)
        prod_data = [['Peringkat', 'Produk', 'Terjual', 'Pendapatan']]
        for idx, p in enumerate(top_products, 1):
            rank_label = f'#{idx}'
            prod_data.append([
                rank_label,
                p['product_name'],
                str(p['total_sold']),
                _format_rupiah(p['total_revenue']),
            ])
        prod_data.append([
            '',
            Paragraph('<b>TOTAL</b>', ParagraphStyle('ProdTotalLbl', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            Paragraph(f'<b>{prod_total_sold}</b>', ParagraphStyle('ProdTotalSold', fontSize=8, leading=10, fontName='Helvetica-Bold')),
            Paragraph(f'<b>{_format_rupiah(prod_total_rev)}</b>', ParagraphStyle('ProdTotalRev', fontSize=8, leading=10, fontName='Helvetica-Bold')),
        ])
        t_prod = Table(prod_data, colWidths=[50, 170, 50, 80])
        t_prod.setStyle(make_table_style(has_total=True))
        elements.append(t_prod)
        elements.append(Spacer(1, 8*mm))

        # ═══════════════════════════════════════════════════════════════════
        # 7.  FOOTER NOTE
        # ═══════════════════════════════════════════════════════════════════
        elements.append(Paragraph(
            f'Laporan ini digenerate secara otomatis oleh Warungio Seller Dashboard '
            f'pada {dt.now().strftime("%d %B %Y pukul %H:%M")} WIB. '
            f'Data bersumber dari transaksi dan aktivitas toko {store_name}.',
            s_footer
        ))

        # Build
        doc.build(elements)
        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="laporan-{store_name.lower().replace(" ", "-")}.pdf"'
        )
        response['Content-Length'] = len(pdf_bytes)
        return response
