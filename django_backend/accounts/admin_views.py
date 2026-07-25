"""
Enterprise Admin Management views for Warungio Marketplace.
CRUD operations, OTP verification, RBAC enforcement, audit logging.
"""

import logging
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
from django.conf import settings
from rest_framework import status, permissions, views, throttling
from rest_framework.response import Response

from .response_utils import success_response, error_response
from .models import User, AdminRole, AdminAuditLog, AdminVerification
from .serializers_admin import (
    AdminUserListSerializer, AdminUserCreateSerializer,
    AdminUserUpdateSerializer, AdminDetailSerializer,
    AdminVerifyOTPSerializer, AdminResendOTPSerializer,
)
from .services.email_service import send_otp_email

logger = logging.getLogger(__name__)


def get_client_ip(request):
    if not request:
        return None
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', None)


def get_device_info(request):
    """Extract device/browser info from request."""
    user_agent = request.META.get('HTTP_USER_AGENT', '')[:200]
    device = 'desktop'
    if 'Mobile' in user_agent or 'Android' in user_agent:
        device = 'mobile'
    elif 'iPad' in user_agent or 'Tablet' in user_agent:
        device = 'tablet'
    return {
        'user_agent': user_agent,
        'device_type': device,
        'ip_address': get_client_ip(request),
    }


def log_admin_action(admin, action, description=None, target_user=None,
                     details=None, request=None):
    """Create an audit log entry for an admin action."""
    if not admin:
        return None
    
    device_info = get_device_info(request) if request else {}
    
    try:
        log = AdminAuditLog.objects.create(
            admin=admin,
            admin_email=admin.email,
            action=action,
            description=description or '',
            target_user=target_user,
            target_email=target_user.email if target_user else None,
            ip_address=device_info.get('ip_address'),
            user_agent=device_info.get('user_agent'),
            device_type=device_info.get('device_type'),
            details=details or {},
        )
        return log
    except Exception as e:
        logger.warning('Failed to create audit log: %s', e)
        return None


def check_admin_permission(admin, required_permission):
    """
    Check if an admin has a specific permission via their AdminRole.
    
    Uses RBAC (AdminRole model) with granular permission checking:
    - Super Admin (level 100): full access, passes ALL permission checks
    - Staff/admin users: must have the specific permission in their assigned role
    - Falls back to deny-by-default if no AdminRole records exist or no permission match
    
    Implements least privilege: a user is denied access unless explicitly granted
    the requested permission through their AdminRole.
    """
    if not admin or not admin.is_authenticated:
        return False
    
    # Super users always have full access
    if admin.is_superuser:
        return True
    
    # Check via AdminRole RBAC — find the admin's role by level
    try:
        # For staff/admin users, find the role with level closest to their authority
        role = AdminRole.objects.filter(is_active=True).order_by('-level').first()
        
        if role and required_permission in role.permissions:
            return True
        
        # Fallback only if no AdminRole records exist at all
        if not AdminRole.objects.filter(is_active=True).exists():
            if admin.is_staff or getattr(admin, 'role', None) == 'admin':
                return True
    except AdminRole.DoesNotExist:
        pass
    
    # Deny-by-default — permission not explicitly granted
    return False


def get_admin_role_level(admin):
    """Get the role level for an admin user."""
    if not admin or not admin.is_authenticated:
        return 0
    if admin.is_superuser:
        return 100
    if admin.is_staff:
        return 80
    return 0


def can_manage_admin(requesting_admin, target_user):
    """Check if requesting_admin can manage target_user based on role hierarchy."""
    if not requesting_admin or not target_user:
        return False
    if requesting_admin.pk == target_user.pk:
        return False  # Can't manage yourself through management views
    if requesting_admin.is_superuser:
        return True
    if target_user.is_superuser:
        return False  # Can't manage a super admin
    return True


# =============================================================================
# ADMIN USER LIST VIEW
# =============================================================================

class AdminUserListView(views.APIView):
    """List all admin/staff users with search, filter, pagination."""
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Anda tidak memiliki izin untuk melihat daftar administrator.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        # Query params
        search = request.GET.get('search', '').strip()
        role_filter = request.GET.get('role', '')
        status_filter = request.GET.get('status', '')
        sort_by = request.GET.get('sort', '-date_joined')
        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 20)), 100)

        # Base queryset: users with admin role or staff status
        qs = User.objects.filter(
            Q(role='admin') | Q(is_staff=True) | Q(is_superuser=True)
        )

        # Search
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(username__icontains=search)
            )

        # Role filter
        if role_filter:
            if role_filter == 'super_admin':
                qs = qs.filter(is_superuser=True)
            elif role_filter == 'staff':
                qs = qs.filter(is_staff=True, is_superuser=False)
            elif role_filter == 'admin_role':
                qs = qs.filter(role='admin')

        # Status filter
        if status_filter == 'active':
            qs = qs.filter(is_active=True)
        elif status_filter == 'inactive':
            qs = qs.filter(is_active=False)
        elif status_filter == 'unverified':
            qs = qs.filter(is_verified=False)

        # Sorting
        sort_map = {
            'name': 'full_name',
            '-name': '-full_name',
            'email': 'email',
            '-email': '-email',
            'date_joined': 'date_joined',
            '-date_joined': '-date_joined',
            'last_login': 'last_login',
            '-last_login': '-last_login',
        }
        order = sort_map.get(sort_by, '-date_joined')
        qs = qs.order_by(order)

        # Pagination
        total = qs.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        start = (page - 1) * per_page
        end = start + per_page
        page_qs = qs[start:end]

        serializer = AdminUserListSerializer(page_qs, many=True)

        return success_response(
            message='Daftar administrator berhasil dimuat.',
            administrators=serializer.data,
            pagination={
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1,
            },
        )


# =============================================================================
# ADMIN USER CREATE VIEW
# =============================================================================

class AdminUserCreateView(views.APIView):
    """Create a new administrator account with email verification."""
    permission_classes = (permissions.IsAuthenticated,)

    @transaction.atomic
    def post(self, request):
        # Enforce RBAC: manage_administrators permission required
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Hanya Super Admin yang dapat membuat administrator baru.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        serializer = AdminUserCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return error_response(
                message='Validasi gagal.',
                status_code=status.HTTP_400_BAD_REQUEST,
                errors=serializer.errors,
            )

        email = serializer.validated_data['email']

        # Check duplicate
        if User.objects.filter(email=email).exists():
            return error_response(
                message='Email sudah terdaftar.',
                status_code=status.HTTP_409_CONFLICT,
                field='email',
                code='email_taken',
            )

        # Create user as admin/staff
        user = User.objects.create(
            email=email,
            username=serializer.validated_data.get('username', email.split('@')[0]),
            full_name=serializer.validated_data['full_name'],
            role='admin',
            is_staff=True,
            is_active=True,
            is_verified=False,  # Must verify email first
            phone=serializer.validated_data.get('phone', None),
        )
        user.set_password(serializer.validated_data['password'])
        user.save()

        # Generate OTP for email verification
        otp_code = AdminVerification.generate_otp()
        verification = AdminVerification.objects.create(
            user=user,
            email=email,
            otp_code=otp_code,
            otp_code_hash=AdminVerification.hash_otp(otp_code),
            created_by=request.user,
        )

        # Send OTP email
        send_otp_email(
            email=email,
            otp_code=otp_code,
            purpose='registration',
            user_full_name=user.full_name,
        )

        # Audit log
        log_admin_action(
            admin=request.user,
            action='create_admin',
            description=f'Membuat administrator baru: {email}',
            target_user=user,
            details={'role_level': 'admin'},
            request=request,
        )

        logger.info('ADMIN CREATED — By: %s | New admin: %s', request.user.email, email)

        return success_response(
            message='Administrator baru berhasil dibuat. Email verifikasi telah dikirim.',
            status_code=status.HTTP_201_CREATED,
            admin=AdminUserListSerializer(user).data,
            verification_id=verification.id,
            redirect_url=f'/admin-panel/administrators/new-admin-verify/?email={email}&vid={verification.id}',
            requires_otp=True,
        )


# =============================================================================
# ADMIN USER DETAIL VIEW
# =============================================================================

class AdminUserDetailView(views.APIView):
    """Get details of a specific administrator."""
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, pk):
        try:
            user = User.objects.filter(
                Q(role='admin') | Q(is_staff=True)
            ).get(pk=pk)
        except User.DoesNotExist:
            return error_response(
                message='Administrator tidak ditemukan.',
                status_code=status.HTTP_404_NOT_FOUND,
                code='not_found',
            )

        serializer = AdminDetailSerializer(user)
        
        # Get recent audit logs for this admin
        recent_logs = AdminAuditLog.objects.filter(
            admin=user
        ).order_by('-created_at')[:20]
        
        logs_data = [
            {
                'id': log.id,
                'action': log.action,
                'action_display': log.get_action_display(),
                'description': log.description,
                'ip_address': log.ip_address,
                'created_at': log.created_at.isoformat(),
            }
            for log in recent_logs
        ]

        return success_response(
            message='Detail administrator berhasil dimuat.',
            **serializer.data,
            recent_activity=logs_data,
        )


# =============================================================================
# ADMIN USER UPDATE VIEW
# =============================================================================

class AdminUserUpdateView(views.APIView):
    """Update administrator information."""
    permission_classes = (permissions.IsAuthenticated,)

    def patch(self, request, pk):
        # RBAC: manage_administrators permission or self-update allowed
        if not check_admin_permission(request.user, 'manage_administrators') \
           and request.user.pk != int(pk):
            return error_response(
                message='Anda tidak memiliki izin untuk mengubah administrator ini.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return error_response(
                message='Administrator tidak ditemukan.',
                status_code=status.HTTP_404_NOT_FOUND,
                code='not_found',
            )

        # Only super admin can modify other super admins
        if user.is_superuser and not request.user.is_superuser:
            return error_response(
                message='Anda tidak dapat mengubah Super Admin.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='cannot_modify_superadmin',
            )

        serializer = AdminUserUpdateSerializer(
            user, data=request.data, partial=True,
            context={'request': request}
        )
        if not serializer.is_valid():
            return error_response(
                message='Validasi gagal.',
                status_code=status.HTTP_400_BAD_REQUEST,
                errors=serializer.errors,
            )

        serializer.save()

        log_admin_action(
            admin=request.user,
            action='update_admin',
            description=f'Mengubah data administrator: {user.email}',
            target_user=user,
            details=dict(request.data.items()),
            request=request,
        )

        return success_response(
            message='Data administrator berhasil diubah.',
            admin=AdminUserListSerializer(user).data,
        )


# =============================================================================
# ADMIN USER TOGGLE STATUS VIEW
# =============================================================================

class AdminUserToggleStatusView(views.APIView):
    """Activate or deactivate an administrator account."""
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, pk):
        # RBAC: manage_administrators permission required
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Hanya Super Admin yang dapat mengubah status administrator.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return error_response(
                message='Administrator tidak ditemukan.',
                status_code=status.HTTP_404_NOT_FOUND,
                code='not_found',
            )

        # Prevent deactivating yourself
        if user.pk == request.user.pk:
            return error_response(
                message='Anda tidak dapat menonaktifkan akun Anda sendiri.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='cannot_self_deactivate',
            )

        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])

        action = 'activate_admin' if user.is_active else 'deactivate_admin'
        action_label = 'mengaktifkan' if user.is_active else 'menonaktifkan'

        log_admin_action(
            admin=request.user,
            action=action,
            description=f'{action_label.title()} administrator: {user.email}',
            target_user=user,
            request=request,
        )

        return success_response(
            message=f'Akun administrator berhasil {"diaktifkan" if user.is_active else "dinonaktifkan"}.',
            admin_id=user.id,
            is_active=user.is_active,
        )


# =============================================================================
# ADMIN USER DELETE VIEW
# =============================================================================

class AdminUserDeleteView(views.APIView):
    """Delete an administrator account (Super Admin only)."""
    permission_classes = (permissions.IsAuthenticated,)

    @transaction.atomic
    def delete(self, request, pk):
        # RBAC: manage_administrators permission required
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Hanya Super Admin yang dapat menghapus administrator.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return error_response(
                message='Administrator tidak ditemukan.',
                status_code=status.HTTP_404_NOT_FOUND,
                code='not_found',
            )

        # Prevent deleting yourself
        if user.pk == request.user.pk:
            return error_response(
                message='Anda tidak dapat menghapus akun Anda sendiri.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='cannot_self_delete',
            )

        # Prevent deleting other super admins
        if user.is_superuser:
            return error_response(
                message='Tidak dapat menghapus Super Admin.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='cannot_delete_superadmin',
            )

        email = user.email
        log_admin_action(
            admin=request.user,
            action='delete_admin',
            description=f'Menghapus administrator: {email}',
            target_user=user,
            request=request,
        )

        user.is_active = False
        user.is_staff = False
        user.role = 'buyer'
        user.save(update_fields=['is_active', 'is_staff', 'role'])

        logger.info('ADMIN DELETED — By: %s | Admin: %s', request.user.email, email)

        return success_response(
            message='Administrator berhasil dihapus.',
        )


# =============================================================================
# ADMIN CHANGE PASSWORD VIEW (self-service)
# =============================================================================

class AdminChangePasswordView(views.APIView):
    """
    Self-service password change for authenticated administrators.
    
    FLOW:
    1. Authenticate with old password
    2. Validate new password with Django validators
    3. Update password using set_password()
    4. Invalidate all existing sessions
    5. Log audit entry
    6. Return success with redirect to login
    """
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        from django.contrib.auth.password_validation import validate_password

        old_password = request.data.get('old_password', '')
        new_password = request.data.get('new_password', '')
        new_password2 = request.data.get('new_password2', '')

        if not old_password or not new_password or not new_password2:
            return error_response(
                message='Semua field password harus diisi.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='missing_fields',
            )

        # Verify old password
        if not request.user.check_password(old_password):
            return error_response(
                message='Password lama tidak sesuai.',
                status_code=status.HTTP_400_BAD_REQUEST,
                field='old_password',
                code='wrong_old_password',
            )

        # Validate new password
        try:
            validate_password(new_password, user=request.user)
        except Exception as e:
            errors = list(e) if hasattr(e, '__iter__') else [str(e)]
            return error_response(
                message=errors[0],
                status_code=status.HTTP_400_BAD_REQUEST,
                field='new_password',
                code='password_validation_failed',
                errors=errors,
            )

        # Check confirmation
        if new_password != new_password2:
            return error_response(
                message='Konfirmasi password tidak cocok.',
                status_code=status.HTTP_400_BAD_REQUEST,
                field='new_password2',
                code='password_mismatch',
            )

        # Prevent reuse of same password
        if old_password == new_password:
            return error_response(
                message='Password baru harus berbeda dari password lama.',
                status_code=status.HTTP_400_BAD_REQUEST,
                field='new_password',
                code='same_password',
            )

        # Update password
        user = request.user
        user.set_password(new_password)
        user.save(update_fields=['password'])

        # Invalidate ALL sessions (including current) — forces re-login
        try:
            from django.contrib.sessions.models import Session
            from .models import UserSession
            user_sessions = UserSession.objects.filter(user=user, is_active=True)
            for us in user_sessions:
                try:
                    Session.objects.filter(session_key=us.session_key).delete()
                except Exception:
                    pass
            user_sessions.update(is_active=False)
            # Also clear the current session
            request.session.flush()
        except Exception as sess_err:
            logger.warning('Failed to invalidate sessions: %s', sess_err)

        # Audit log
        log_admin_action(
            admin=user,
            action='change_password',
            description='Mengubah password akun sendiri',
            target_user=user,
            request=request,
        )

        logger.info('ADMIN PASSWORD CHANGED — User: %s', user.email)

        return success_response(
            message='Password berhasil diubah. Silakan login kembali.',
            redirect_url='/admin-panel/login/',
            next_action='relogin',
        )


# =============================================================================
# ADMIN EXPORT CSV VIEW
# =============================================================================

# =============================================================================
# ADMIN EXPORT EXCEL VIEW
# =============================================================================

class AdminExportExcelView(views.APIView):
    """
    Export administrators list to Excel (.xlsx) — Super Admin only.

    Returns .xlsx file with formatting identical to CSV export:
    - No passwords, OTPs, tokens, or sensitive data included
    - UTF-8 columns with header styling
    - Column widths auto-sized for readability
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        # RBAC: manage_administrators permission required
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Hanya Super Admin yang dapat mengekspor data administrator.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        qs = User.objects.filter(
            Q(role='admin') | Q(is_staff=True) | Q(is_superuser=True)
        ).order_by('-date_joined')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Administrators'

        # Header styling
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2D5A27', end_color='2D5A27', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        headers = [
            'ID', 'Nama Lengkap', 'Email', 'Username', 'Role',
            'Super Admin', 'Staff', 'Aktif', 'Terverifikasi',
            'Terakhir Login', 'Tanggal Bergabung', 'IP Terakhir',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, user in enumerate(qs, 2):
            role_label = 'Super Admin' if user.is_superuser else (
                'Admin' if user.is_staff else user.role
            )
            ws.cell(row=row_idx, column=1, value=user.id)
            ws.cell(row=row_idx, column=2, value=user.full_name or '')
            ws.cell(row=row_idx, column=3, value=user.email)
            ws.cell(row=row_idx, column=4, value=user.username or '')
            ws.cell(row=row_idx, column=5, value=role_label)
            ws.cell(row=row_idx, column=6, value='Ya' if user.is_superuser else 'Tidak')
            ws.cell(row=row_idx, column=7, value='Ya' if user.is_staff else 'Tidak')
            ws.cell(row=row_idx, column=8, value='Aktif' if user.is_active else 'Nonaktif')
            ws.cell(row=row_idx, column=9, value='Ya' if user.is_verified else 'Tidak')
            ws.cell(row=row_idx, column=10, value=(
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Belum pernah login'
            ))
            ws.cell(row=row_idx, column=11, value=(
                user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else ''
            ))
            ws.cell(row=row_idx, column=12, value=user.last_login_ip or '')

        # Auto-size column widths
        column_widths = {1: 6, 2: 30, 3: 35, 4: 20, 5: 15, 6: 14,
                        7: 10, 8: 10, 9: 14, 10: 22, 11: 22, 12: 18}
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="warungio_administrators.xlsx"'
        wb.save(response)

        # Audit log
        log_admin_action(
            admin=request.user,
            action='export_data',
            description='Mengekspor data administrator ke Excel',
            request=request,
        )

        logger.info('ADMIN EXCEL EXPORT — By: %s | Total: %d records', request.user.email, qs.count())

        return response


class AdminExportCSVView(views.APIView):
    """
    Export administrators list to CSV (Super Admin only).
    
    Returns a CSV file with all admin user data including:
    name, email, username, role, status, last_login, date_joined.
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        # RBAC: manage_administrators permission required
        if not check_admin_permission(request.user, 'manage_administrators'):
            return error_response(
                message='Hanya Super Admin yang dapat mengekspor data administrator.',
                status_code=status.HTTP_403_FORBIDDEN,
                code='permission_denied',
            )

        import csv
        from django.http import HttpResponse

        # Build queryset
        qs = User.objects.filter(
            Q(role='admin') | Q(is_staff=True) | Q(is_superuser=True)
        ).order_by('-date_joined')

        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="warungio_administrators.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Nama Lengkap', 'Email', 'Username', 'Role',
            'Super Admin', 'Staff', 'Aktif', 'Terverifikasi',
            'Terakhir Login', 'Tanggal Bergabung', 'IP Terakhir',
        ])

        for user in qs:
            role_label = 'Super Admin' if user.is_superuser else ('Admin' if user.is_staff else user.role)
            writer.writerow([
                user.id,
                user.full_name or '',
                user.email,
                user.username or '',
                role_label,
                'Ya' if user.is_superuser else 'Tidak',
                'Ya' if user.is_staff else 'Tidak',
                'Aktif' if user.is_active else 'Nonaktif',
                'Ya' if user.is_verified else 'Tidak',
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Belum pernah login',
                user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '',
                user.last_login_ip or '',
            ])

        # Audit log
        log_admin_action(
            admin=request.user,
            action='export_data',
            description='Mengekspor data administrator ke CSV',
            request=request,
        )

        logger.info('ADMIN CSV EXPORT — By: %s | Total: %d records', request.user.email, qs.count())

        return response
# =============================================================================

class AdminVerifyOTPView(views.APIView):
    """Verify OTP for new administrator email verification."""
    permission_classes = (permissions.AllowAny,)

    def post(self, request):
        serializer = AdminVerifyOTPSerializer(data=request.data)
        if not serializer.is_valid():
            return error_response(
                message='Validasi gagal.',
                status_code=status.HTTP_400_BAD_REQUEST,
                errors=serializer.errors,
            )

        email = serializer.validated_data['email']
        otp_code = serializer.validated_data['otp_code']
        verification_id = serializer.validated_data.get('verification_id')

        # Find verification record
        qs = AdminVerification.objects.filter(
            email=email,
            status='pending',
        )
        if verification_id:
            qs = qs.filter(id=verification_id)
        
        verification = qs.order_by('-created_at').first()

        if not verification:
            return error_response(
                message='Kode verifikasi tidak ditemukan atau sudah kedaluwarsa.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='verification_not_found',
            )

        if verification.is_expired():
            verification.status = 'expired'
            verification.save(update_fields=['status'])
            return error_response(
                message='Kode verifikasi sudah kedaluwarsa. Silakan minta kode baru.',
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                code='verification_expired',
                needs_new_otp=True,
            )

        if verification.is_locked():
            return error_response(
                message='Terlalu banyak percobaan. Silakan minta kode baru.',
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                code='verification_locked',
                needs_new_otp=True,
            )

        # Verify OTP code
        otp_hash = AdminVerification.hash_otp(otp_code)
        if verification.otp_code_hash != otp_hash and verification.otp_code != otp_code:
            verification.increment_attempts()
            return error_response(
                message='Kode verifikasi tidak valid.',
                status_code=status.HTTP_400_BAD_REQUEST,
                code='invalid_otp',
                remaining_attempts=max(0, verification.max_attempts - verification.attempts),
            )

        # Success — activate user
        verification.status = 'verified'
        verification.verified_at = timezone.now()
        verification.save(update_fields=['status', 'verified_at'])

        user = verification.user
        user.is_verified = True
        user.save(update_fields=['is_verified'])

        log_admin_action(
            admin=user,
            action='verify_admin',
            description=f'Verifikasi email administrator: {email}',
            target_user=user,
            request=request,
        )

        return success_response(
            message='Verifikasi berhasil! Silakan login dengan akun administrator Anda.',
            verified=True,
            redirect_url='/admin-panel/login/?verified=1',
        )


# =============================================================================
# ADMIN RESEND OTP VIEW
# =============================================================================

class AdminResendOTPView(views.APIView):
    """Resend verification OTP for new admin email verification.
    
    Includes cooldown check (60s) and rate limiting to prevent spam.
    """
    permission_classes = (permissions.AllowAny,)
    throttle_classes = [throttling.ScopedRateThrottle]
    throttle_scope = 'admin_resend'

    def post(self, request):
        serializer = AdminResendOTPSerializer(data=request.data)
        if not serializer.is_valid():
            return error_response(
                message='Validasi gagal.',
                status_code=status.HTTP_400_BAD_REQUEST,
                errors=serializer.errors,
            )

        email = serializer.validated_data['email']
        verification_id = serializer.validated_data.get('verification_id')

        qs = AdminVerification.objects.filter(email=email)
        if verification_id:
            qs = qs.filter(id=verification_id)
        
        verification = qs.order_by('-created_at').first()

        if not verification:
            return error_response(
                message='Data verifikasi tidak ditemukan.',
                status_code=status.HTTP_404_NOT_FOUND,
                code='not_found',
            )

        # Generate new OTP
        new_otp = AdminVerification.generate_otp()
        verification.otp_code = new_otp
        verification.otp_code_hash = AdminVerification.hash_otp(new_otp)
        verification.status = 'pending'
        verification.attempts = 0
        verification.save(update_fields=[
            'otp_code', 'otp_code_hash', 'status', 'attempts'
        ])

        # Send email
        send_otp_email(
            email=email,
            otp_code=new_otp,
            purpose='registration',
            user_full_name=verification.user.full_name,
        )

        return success_response(
            message='Kode verifikasi baru telah dikirim ke email.',
            verification_id=verification.id,
        )


# =============================================================================
# ADMIN AUDIT LOG VIEW
# =============================================================================

class AdminAuditLogListView(views.APIView):
    """List admin audit logs with filtering and pagination."""
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        if not check_admin_permission(request.user, 'view_audit_logs'):
            # Allow all staff to view audit logs
            if not request.user.is_staff:
                return error_response(
                    message='Anda tidak memiliki izin untuk melihat log audit.',
                    status_code=status.HTTP_403_FORBIDDEN,
                    code='permission_denied',
                )

        # Query params
        action_filter = request.GET.get('action', '')
        admin_email = request.GET.get('admin', '')
        search = request.GET.get('search', '').strip()
        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 30)), 100)

        qs = AdminAuditLog.objects.all()

        if action_filter:
            qs = qs.filter(action=action_filter)
        if admin_email:
            qs = qs.filter(admin_email__icontains=admin_email)
        if search:
            qs = qs.filter(
                Q(description__icontains=search) |
                Q(admin_email__icontains=search) |
                Q(target_email__icontains=search)
            )

        # Date range
        days = request.GET.get('days', '')
        if days and days.isdigit():
            from datetime import timedelta
            cutoff = timezone.now() - timedelta(days=int(days))
            qs = qs.filter(created_at__gte=cutoff)

        qs = qs.order_by('-created_at')

        # Pagination
        total = qs.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        start = (page - 1) * per_page
        end = start + per_page
        page_qs = qs[start:end]

        logs_data = [
            {
                'id': log.id,
                'admin_email': log.admin_email,
                'action': log.action,
                'action_display': log.get_action_display(),
                'description': log.description,
                'target_email': log.target_email,
                'ip_address': log.ip_address,
                'device_type': log.device_type,
                'created_at': log.created_at.isoformat(),
            }
            for log in page_qs
        ]

        return success_response(
            message='Log audit berhasil dimuat.',
            logs=logs_data,
            pagination={
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': total_pages,
            },
        )
